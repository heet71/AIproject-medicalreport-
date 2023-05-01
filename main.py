from flask import Flask, render_template, request, redirect, url_for, send_file #flask are used to connect the web page or web server
import PyPDF2   #PyPDF2 Library used to Extract text from upload pdf
import openpyxl #get information to store excel sheet
import os

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['DOWNLOAD_FOLDER'] = 'downloads'


@app.route('/')
def index():
    return render_template('index.html')    #access the web page


import spacy

# Load the SpaCy model
nlp = spacy.load('en_core_web_sm')

@app.route('/upload', methods=['POST'])     # upload pdf file code
def upload():
    # Get the uploaded PDF file
    pdf_file = request.files['pdf']
    pdf_file_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_file.filename)
    pdf_file.save(pdf_file_path)

    # Create a PdfReader object
    pdf_reader = PyPDF2.PdfReader(pdf_file_path)

    # Create an empty list to store the extracted text
    extracted_text = []

    # Loop through each page of the PDF file and extract the text
    for page_num in range(len(pdf_reader.pages)):
        # do something with each page
        page = pdf_reader.pages[page_num]
        # example: extract text from page
        text = page.extract_text()
        print(text)
        extracted_text.append(text)  # Append the extracted text to the list

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    # Add column headers
    worksheet.cell(row=1, column=1, value='Name')
    worksheet.cell(row=1, column=2, value='Date of Birth')
    worksheet.cell(row=1, column=3, value='City')
    worksheet.cell(row=1, column=4, value='Doctor Name')

    # Loop through each item in the extracted_text list and write it to the Excel worksheet
    for row_num, text in enumerate(extracted_text, start=1):  # Start from row 2 to skip header row
        # Process the text with SpaCy
        doc = nlp(text)
        name = ''
        dob = ''
        city = ''
        doctor_name = ''

        # Extract the relevant information from SpaCy's named entity recognition (NER)
        for ent in doc.ents:
            if ent.label_ == 'PERSON' and not name:
                name = ent.text
            elif ent.label_ == 'DATE' and not dob:
                dob = ent.text
            elif ent.label_ == 'GPE' and not city:
                city = ent.text
            elif ent.label_ == 'PERSON' and not doctor_name:
                doctor_name = ent.text

            # Write data to corresponding columns
            worksheet.cell(row=row_num, column=1, value=name)
            worksheet.cell(row=row_num, column=2, value=dob)
            worksheet.cell(row=row_num, column=3, value=city)
            worksheet.cell(row=row_num, column=4, value=doctor_name)

    # Save the Excel workbook
    excel_file_path = os.path.join(app.config['DOWNLOAD_FOLDER'], 'extracted_text.xlsx')
    workbook.save(excel_file_path)

    return redirect(url_for('index', extracted_text=extracted_text))

@app.route('/download_excel')
def download_excel():
    excel_file_path = os.path.join(app.config['DOWNLOAD_FOLDER'], 'extracted_text.xlsx')
    return send_file(excel_file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
